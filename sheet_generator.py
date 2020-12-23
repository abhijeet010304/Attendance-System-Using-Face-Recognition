import openpyxl
from datetime import date, datetime
import calendar

def findDay(date): 
    born = datetime.strptime(date, '%d %m %Y').weekday() 
    return (calendar.day_name[born]) 

def generate_sheet(number_of_students):
	attendance_sheet = openpyxl.load_workbook('Attendance.xlsx')
	today = date.today()
	month_string = today.strftime("%B")

	if today.strftime("%d") == 1 or month_string not in attendance_sheet.sheetnames:
		attendance_sheet.create_sheet(month_string)

	sheet = attendance_sheet[month_string]


	now = datetime.now()
	number_of_days = calendar.monthrange(now.year, now.month)[1]


	for i in range(2, number_of_students+2):
		roll = str(i-1)
		while len(roll) != 3:
			roll = '0' + roll
		sheet.cell(row=1, column=i).value = 'BT18CSE' + roll

	total_working_days = 0
	rownum = 2
	for i in range(1, number_of_days+1):
		date_today = today.strftime("%d/%m/%Y")
		date_string = str(i) + " " + date_today.split('/')[1] + " " + date_today.split('/')[2]
		day_string = findDay(date_string)
		if day_string != 'Sunday':
			sheet.cell(row = rownum, column = 1).value = str(i) + " (" + day_string + ")"
			rownum += 1
			total_working_days += 1

	sheet.cell(row = 33, column = 1).value = "TOTAL"
	sheet.cell(row = 34, column = 1).value = total_working_days
	sheet.cell(row = 36, column = 1).value = 'PERCENTAGE'
	for i in range(2, number_of_students+2):
		col = openpyxl.utils.get_column_letter(i)
		begin = col+'2'
		end = col+str(total_working_days+1)
		string = '=SUM(' + begin + ":" + end + ")"
		sheet.cell(row=34, column=i).value = string

		div_cell = col+'34'
		string = '='+div_cell + '*' + str(100/total_working_days)
		sheet.cell(row=36, column=i).value = string



	for i in range(2, number_of_students+2):
		for j in range(2, total_working_days+2):
			sheet.cell(row = j, column = i).value = 0

	attendance_sheet.save('Attendance.xlsx')
